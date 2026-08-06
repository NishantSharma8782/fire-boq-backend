"""
Export Service - Generates PDF and Excel reports from project BOQ data.
Architectural Obsidian & Flame Edition - High-Impact PDF Report.
"""
import io
import csv
from datetime import datetime
from typing import List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable, KeepTogether
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfgen import canvas


class NumberedCanvas(canvas.Canvas):
    """Two-pass canvas to dynamically compute total pages and draw running headers/footers."""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count):
        self.saveState()
        
        # ── Running Top Header (Pages 2+) ──
        if self._pageNumber > 1:
            self.setFont("Helvetica-Bold", 8)
            self.setFillColor(colors.HexColor("#0F172A"))
            self.drawString(1.2 * cm, 28.5 * cm, "FIRE ENGINEERING TAKEOFF — BILL OF QUANTITIES (BOQ)")
            self.setFont("Helvetica", 7.5)
            self.setFillColor(colors.HexColor("#64748B"))
            self.drawRightString(19.8 * cm, 28.5 * cm, "OFFICIAL ENGINEERING REPORT")
            self.setStrokeColor(colors.HexColor("#EA580C"))
            self.setLineWidth(1.8)
            self.line(1.2 * cm, 28.3 * cm, 19.8 * cm, 28.3 * cm)

        # ── Running Footer (All Pages) ──
        self.setFont("Helvetica-Bold", 7.5)
        self.setFillColor(colors.HexColor("#0F172A"))
        self.drawString(1.2 * cm, 0.8 * cm, "STRICTLY CONFIDENTIAL")
        self.setFont("Helvetica", 7.5)
        self.setFillColor(colors.HexColor("#64748B"))
        self.drawString(4.5 * cm, 0.8 * cm, "• AI Fire Safety BOQ Engine • Subject to Final PE Verification")
        page_str = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(19.8 * cm, 0.8 * cm, page_str)
        self.setStrokeColor(colors.HexColor("#CBD5E1"))
        self.setLineWidth(0.6)
        self.line(1.2 * cm, 1.1 * cm, 19.8 * cm, 1.1 * cm)
        
        self.restoreState()


def generate_pdf(project: dict, building_data: dict, recommendations: dict, boq_sections: list, standard: str = "NBC") -> bytes:
    """Generate an Executive Classic Architectural & Engineering PDF BOQ Report."""
    is_nfpa = standard.upper() == "NFPA"
    std_display = "NFPA 72 / 13 / 14 / 10 COMPLIANT" if is_nfpa else "NBC 2016 / IS 2189 COMPLIANT"
    std_notes = (
        "NFPA 72 (Fire Alarm), NFPA 13 (Sprinklers), NFPA 14 (Standpipes), NFPA 10 (Extinguishers)"
        if is_nfpa
        else "NBC 2016 Part 4, IS 2189:2008, IS 15105:2002, IS 3844:1989, IS 2190:2010"
    )
    
    buffer = io.BytesIO()
    
    # Document dimensions: A4 portrait (21.0cm x 29.7cm)
    # Printable area: width = 18.6cm (527.24 pt)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.2 * cm,
        leftMargin=1.2 * cm,
        topMargin=1.4 * cm,
        bottomMargin=1.4 * cm,
        title=f"Fire BOQ - {project.get('project_name', 'Report')}",
    )

    styles = getSampleStyleSheet()
    story = []

    # ── CLASSIC EXECUTIVE PALETTE ────────────────────────────────────────────────
    COLOR_NAVY_MAIN   = colors.HexColor("#0B132A")  # Deep Executive Navy
    COLOR_NAVY_HEADER = colors.HexColor("#1C2541")  # Slate Navy Header
    COLOR_GOLD_ACCENT = colors.HexColor("#D4AF37")  # Classic Gold
    COLOR_AMBER_TEXT  = colors.HexColor("#D97706")  # Rich Amber Text
    COLOR_FLAME_RED   = colors.HexColor("#DC2626")  # Fire Red Accent
    COLOR_TEXT_DARK   = colors.HexColor("#0F172A")  # Deep Charcoal Main Body
    COLOR_TEXT_MUTED  = colors.HexColor("#475569")  # Slate Muted Subtext
    COLOR_BG_LIGHT    = colors.HexColor("#F8FAFC")  # Off-white Card Tint
    COLOR_BG_GOLD     = colors.HexColor("#FEFCE8")  # Warm Gold Tint
    COLOR_BG_RED      = colors.HexColor("#FEF2F2")  # Soft Flame Tint
    COLOR_BORDER      = colors.HexColor("#CBD5E1")  # Classic Border Gray
    COLOR_BORDER_GOLD = colors.HexColor("#EAB308")  # Gold Border Accent
    WHITE             = colors.white

    # ── TYPOGRAPHY & PARAGRAPH STYLES ────────────────────────────────────────────
    title_style = ParagraphStyle(
        "DocTitle", parent=styles["Title"],
        fontName="Helvetica-Bold", fontSize=15, textColor=WHITE,
        alignment=TA_LEFT, leading=18
    )
    subtitle_style = ParagraphStyle(
        "DocSubTitle", parent=styles["Normal"],
        fontName="Helvetica-Bold", fontSize=8.5, textColor=COLOR_GOLD_ACCENT,
        alignment=TA_LEFT, leading=11
    )
    header_proj_name = ParagraphStyle(
        "HProjName", parent=styles["Normal"],
        fontName="Helvetica", fontSize=9.5, textColor=colors.HexColor("#E2E8F0"),
        alignment=TA_LEFT, leading=12
    )
    header_badge_style = ParagraphStyle(
        "HBadge", parent=styles["Normal"],
        fontName="Helvetica-Bold", fontSize=8.5, textColor=WHITE,
        alignment=TA_RIGHT, leading=11
    )
    header_badge_sub = ParagraphStyle(
        "HBadgeSub", parent=styles["Normal"],
        fontName="Helvetica", fontSize=7.5, textColor=colors.HexColor("#94A3B8"),
        alignment=TA_RIGHT, leading=9.5
    )

    section_heading = ParagraphStyle(
        "SecHeading", parent=styles["Heading2"],
        fontName="Helvetica-Bold", fontSize=9.5, textColor=WHITE,
        spaceBefore=0, spaceAfter=0, leading=12
    )

    tbl_head_left = ParagraphStyle(
        "TblHeadL", parent=styles["Normal"],
        fontName="Helvetica-Bold", fontSize=8, textColor=WHITE,
        alignment=TA_LEFT, leading=10
    )
    tbl_head_center = ParagraphStyle("TblHeadC", parent=tbl_head_left, alignment=TA_CENTER)
    tbl_head_right  = ParagraphStyle("TblHeadR", parent=tbl_head_left, alignment=TA_RIGHT)

    tbl_sno_style = ParagraphStyle(
        "TblSno", parent=styles["Normal"],
        fontName="Helvetica-Bold", fontSize=8, textColor=COLOR_FLAME_RED,
        alignment=TA_CENTER, leading=10
    )
    tbl_item_style = ParagraphStyle(
        "TblItem", parent=styles["Normal"],
        fontName="Helvetica-Bold", fontSize=8.5, textColor=COLOR_TEXT_DARK,
        leading=11
    )
    tbl_desc_style = ParagraphStyle(
        "TblDesc", parent=styles["Normal"],
        fontName="Helvetica", fontSize=7.8, textColor=COLOR_TEXT_DARK,
        leading=10
    )
    tbl_unit_style = ParagraphStyle(
        "TblUnit", parent=styles["Normal"],
        fontName="Helvetica", fontSize=7.8, textColor=COLOR_TEXT_MUTED,
        alignment=TA_CENTER, leading=10
    )
    tbl_qty_style = ParagraphStyle(
        "TblQty", parent=styles["Normal"],
        fontName="Helvetica-Bold", fontSize=9, textColor=COLOR_AMBER_TEXT,
        alignment=TA_RIGHT, leading=11
    )
    tbl_basis_style = ParagraphStyle(
        "TblBasis", parent=styles["Normal"],
        fontName="Helvetica-Oblique", fontSize=7.2, textColor=COLOR_TEXT_MUTED,
        leading=9.5
    )

    meta_label = ParagraphStyle(
        "MetaLabel", parent=styles["Normal"],
        fontName="Helvetica-Bold", fontSize=7.2, textColor=COLOR_AMBER_TEXT,
        leading=9.5
    )
    meta_val = ParagraphStyle(
        "MetaVal", parent=styles["Normal"],
        fontName="Helvetica", fontSize=8, textColor=COLOR_TEXT_DARK,
        leading=10.5
    )
    meta_val_bold = ParagraphStyle(
        "MetaValBold", parent=styles["Normal"],
        fontName="Helvetica-Bold", fontSize=8.2, textColor=COLOR_NAVY_MAIN,
        leading=10.5
    )

    # Metrics sum
    total_items_count = 0
    total_qty_sum = 0
    for sec in boq_sections:
        for it in sec.get("items", []):
            total_items_count += 1
            try:
                total_qty_sum += float(it.get("quantity", 0))
            except (ValueError, TypeError):
                pass

    # ── 1. EXECUTIVE COVER HEADER BANNER ─────────────────────────────────────────
    header_left = [
        Paragraph("OFFICIAL FIRE ENGINEERING TAKEOFF REPORT", title_style),
        Paragraph(f"BILL OF QUANTITIES (BOQ) • {std_display}", subtitle_style),
        Spacer(1, 2),
        Paragraph(f"<b>Project:</b> {project.get('project_name', 'N/A')}", header_proj_name)
    ]
    header_right = [
        Paragraph(f"PROJECT ID: {project.get('project_id', 'N/A')}", header_badge_style),
        Paragraph(f"DATE: {datetime.now().strftime('%d %b %Y, %H:%M')}", header_badge_sub),
        Paragraph(f"STATUS: AUDITED ENGINEERING BOQ", header_badge_sub),
        Paragraph(f"STANDARD: {standard.upper()}", header_badge_sub),
    ]

    banner_table = Table([[header_left, header_right]], colWidths=[12.6 * cm, 6.0 * cm])
    banner_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), COLOR_NAVY_MAIN),
        ("PADDING", (0, 0), (-1, -1), 12),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LINEBELOW", (0, 0), (-1, -1), 3.5, COLOR_GOLD_ACCENT),
    ]))
    story.append(banner_table)
    story.append(Spacer(1, 8))

    # ── 2. EXECUTIVE METRICS DASHBOARD (3 TILES ROW) ──────────────────────────
    m1_content = [
        Paragraph("PROJECT SPECIFICATION", meta_label),
        Paragraph(f"<b>{project.get('project_name', 'N/A')}</b>", meta_val_bold),
        Paragraph(f"Client: {project.get('client_name', 'N/A')}", meta_val),
        Paragraph(f"Location: {project.get('location', 'N/A')}", meta_val),
        Paragraph(f"Hazard Category: <b>{project.get('hazard_category', 'light').title()}</b>", meta_val),
    ]
    t_m1 = Table([[m1_content]], colWidths=[6.0 * cm])
    t_m1.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), COLOR_BG_GOLD),
        ("BOX", (0, 0), (-1, -1), 0.8, COLOR_BORDER_GOLD),
        ("PADDING", (0, 0), (-1, -1), 7),
    ]))

    m2_content = [
        Paragraph("SPATIAL & BUILDING DATA", ParagraphStyle("M2L", parent=meta_label, textColor=colors.HexColor("#2563EB"))),
        Paragraph(f"<b>{building_data.get('estimated_area', 0):,.0f} m² Total Floor Area</b>", meta_val_bold),
        Paragraph(f"Floors: {building_data.get('floors', 1)} | Rooms: {building_data.get('rooms', 0)}", meta_val),
        Paragraph(f"Corridors: {building_data.get('corridors', 0)} | Exits: {building_data.get('exits', 0)}", meta_val),
        Paragraph(f"Staircases: {building_data.get('stairs', 0)} | Height: {building_data.get('ceiling_height', 3)}m", meta_val),
    ]
    t_m2 = Table([[m2_content]], colWidths=[6.0 * cm])
    t_m2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), COLOR_BG_LIGHT),
        ("BOX", (0, 0), (-1, -1), 0.8, COLOR_BORDER),
        ("PADDING", (0, 0), (-1, -1), 7),
    ]))

    m3_content = [
        Paragraph("FIRE TAKEOFF SUMMARY", ParagraphStyle("M3L", parent=meta_label, textColor=COLOR_FLAME_RED)),
        Paragraph(f"<b>{int(total_qty_sum):,} Total Devices & Units</b>", meta_val_bold),
        Paragraph(f"Line Items: {total_items_count} across {len(boq_sections)} Sections", meta_val),
        Paragraph(f"Building Type: {project.get('building_type', 'N/A').title()}", meta_val),
        Paragraph(f"Fire Code: {standard.upper()} ({std_display.split(' ')[0]})", meta_val),
    ]
    t_m3 = Table([[m3_content]], colWidths=[6.0 * cm])
    t_m3.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), COLOR_BG_RED),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#FCA5A5")),
        ("PADDING", (0, 0), (-1, -1), 7),
    ]))

    dash_row = Table([[t_m1, t_m2, t_m3]], colWidths=[6.2 * cm, 6.2 * cm, 6.2 * cm])
    dash_row.setStyle(TableStyle([
        ("PADDING", (0, 0), (-1, -1), 0),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(dash_row)
    story.append(Spacer(1, 10))

    # ── 3. BOQ SECTIONS & TABLES ──────────────────────────────────────────────
    col_widths = [1.1 * cm, 3.7 * cm, 6.3 * cm, 1.3 * cm, 1.8 * cm, 4.4 * cm]

    table_header_row = [
        Paragraph("S.No", tbl_head_center),
        Paragraph("Item Specification", tbl_head_left),
        Paragraph("Technical Description", tbl_head_left),
        Paragraph("Unit", tbl_head_center),
        Paragraph("Qty", tbl_head_right),
        Paragraph("Calculation Basis & Code Standard", tbl_head_left),
    ]

    section_accent_colors = {
        "A": colors.HexColor("#DC2626"),  # Flame Red
        "B": colors.HexColor("#2563EB"),  # Royal Blue
        "C": colors.HexColor("#059669"),  # Emerald Green
        "D": colors.HexColor("#D97706"),  # Warm Amber
    }

    for section in boq_sections:
        sec_id = section.get("section_id", "")
        sec_name = section.get("section_name", "").upper()
        accent_c = section_accent_colors.get(sec_id, COLOR_FLAME_RED)
        items = section.get("items", [])

        # Full-width section header banner
        sec_banner_table = Table(
            [[Paragraph(f"SECTION {sec_id}: {sec_name}", section_heading)]],
            colWidths=[18.6 * cm]
        )
        sec_banner_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), COLOR_NAVY_HEADER),
            ("LINELEFT", (0, 0), (0, -1), 5.0, accent_c),
            ("PADDING", (0, 0), (-1, -1), 6.0),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))

        table_data = [table_header_row]
        for item in items:
            sno_val = item.get("sno", "")
            sno_str = f"{int(sno_val):02d}" if str(sno_val).isdigit() else str(sno_val)
            table_data.append([
                Paragraph(sno_str, tbl_sno_style),
                Paragraph(item.get("item", ""), tbl_item_style),
                Paragraph(item.get("description", ""), tbl_desc_style),
                Paragraph(item.get("unit", ""), tbl_unit_style),
                Paragraph(f"{item.get('quantity', 0):,.1f}", tbl_qty_style),
                Paragraph(item.get("calculation_basis", ""), tbl_basis_style),
            ])

        boq_table = Table(table_data, colWidths=col_widths, repeatRows=1)
        boq_table.setStyle(TableStyle([
            # Header Row
            ("BACKGROUND", (0, 0), (-1, 0), COLOR_NAVY_MAIN),
            ("PADDING", (0, 0), (-1, 0), 5),
            ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
            # Data Rows
            ("GRID", (0, 0), (-1, -1), 0.4, COLOR_BORDER),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, COLOR_BG_LIGHT]),
            ("PADDING", (0, 1), (-1, -1), 5),
            ("VALIGN", (0, 1), (-1, -1), "TOP"),
        ]))

        # Keep title banner together with table
        story.append(KeepTogether([sec_banner_table, Spacer(1, 2), boq_table]))
        story.append(Spacer(1, 10))

    # ── 4. OFFICIAL ENGINEERING CERTIFICATE & SIGN-OFF ───────────────────────
    disclaimer_p = Paragraph(
        f"<b>ENGINEERING AUDIT STATEMENT & DISCLAIMER:</b> This Bill of Quantities (BOQ) takeoff was computed by spatial analysis algorithms compliant with {std_notes}. "
        "All device quantities, pipe diameters, conduit runs, and equipment specifications are official engineering estimates. "
        "Review, verification, and stamping by a Registered Chartered Fire Protection Engineer is required prior to commercial tender issuance.",
        ParagraphStyle("AuditP", parent=styles["Normal"], fontName="Helvetica", fontSize=7.2, textColor=COLOR_TEXT_MUTED, leading=9.8)
    )

    stamp_box = [
        Paragraph("<b>CHARTERED PE APPROVAL STAMP</b>", ParagraphStyle("StHead", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=7.5, textColor=COLOR_NAVY_MAIN, alignment=TA_CENTER)),
        Spacer(1, 12),
        Paragraph("_____________________________", ParagraphStyle("Line1", parent=styles["Normal"], fontName="Helvetica", fontSize=8, textColor=COLOR_TEXT_MUTED, alignment=TA_CENTER)),
        Paragraph("Registered Fire Protection Engineer", ParagraphStyle("Line2", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=7, textColor=COLOR_TEXT_MUTED, alignment=TA_CENTER)),
        Paragraph("PE License & Reg. No.", ParagraphStyle("Line3", parent=styles["Normal"], fontName="Helvetica-Oblique", fontSize=6.5, textColor=COLOR_TEXT_MUTED, alignment=TA_CENTER)),
    ]
    t_stamp = Table([[stamp_box]], colWidths=[5.4 * cm])
    t_stamp.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), COLOR_BG_LIGHT),
        ("BOX", (0, 0), (-1, -1), 0.8, COLOR_BORDER_GOLD),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))

    cert_table = Table([[disclaimer_p, t_stamp]], colWidths=[12.8 * cm, 5.8 * cm])
    cert_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), COLOR_BG_LIGHT),
        ("BOX", (0, 0), (-1, -1), 0.8, COLOR_NAVY_HEADER),
        ("PADDING", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(KeepTogether([cert_table]))

    # Build PDF using NumberedCanvas
    doc.build(story, canvasmaker=NumberedCanvas)
    buffer.seek(0)
    return buffer.read()



def generate_excel(project: dict, building_data: dict, recommendations: dict, boq_sections: list, standard: str = "NBC") -> bytes:
    """Generate a formatted Excel BOQ report."""
    std_display = "NFPA 72/13/14/10" if standard.upper() == "NFPA" else "NBC 2016 / IS 2189"
    std_notes   = "NFPA 72, NFPA 13, NFPA 14, NFPA 10" if standard.upper() == "NFPA" else "NBC 2016 Part 4, IS 2189:2008, IS 15105:2002"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fire BOQ"

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 35

    # Style helpers
    def header_style(cell, bg="#2C3E50", fg="FFFFFF", bold=True, size=10):
        cell.fill = PatternFill("solid", fgColor=bg.lstrip("#"))
        cell.font = Font(bold=bold, color=fg, size=size)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def data_style(cell, bold=False, align="left", size=9):
        cell.font = Font(bold=bold, size=size)
        cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # ── TITLE ─────────────────────────────────────────────────────────────────────
    row = 1
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws[f"A{row}"]
    cell.value = "FIRE PROTECTION SYSTEM — BILL OF QUANTITIES"
    header_style(cell, "#C0392B", size=14)
    ws.row_dimensions[row].height = 28

    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws[f"A{row}"]
    cell.value = f"Project: {project.get('project_name', '')}  |  ID: {project.get('project_id', '')}  |  Client: {project.get('client_name', '')}"
    cell.font = Font(bold=True, size=10, color="2C3E50")
    cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[row].height = 18

    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws[f"A{row}"]
    cell.value = (
        f"Location: {project.get('location', '')}  |  Building Type: {project.get('building_type', '').title()}  "
        f"|  Hazard: {project.get('hazard_category', '').title()}  "
        f"|  Standard: {std_display}  "
        f"|  Date: {datetime.now().strftime('%d-%m-%Y')}"
    )
    cell.font = Font(size=9, color="7F8C8D")
    cell.alignment = Alignment(horizontal="center")

    row += 2

    # ── SUMMARY CARDS ─────────────────────────────────────────────────────────────
    summary_headers = ["Total Area (sqm)", "Floors", "Rooms", "Corridors", "Staircases", "Exits"]
    summary_values  = [
        building_data.get("estimated_area", 0),
        building_data.get("floors", 1),
        building_data.get("rooms", 0),
        building_data.get("corridors", 0),
        building_data.get("stairs", 0),
        building_data.get("exits", 0),
    ]

    for col_idx, (h, v) in enumerate(zip(summary_headers, summary_values), start=1):
        col_letter = get_column_letter(col_idx)
        c_h = ws[f"{col_letter}{row}"]
        c_h.value = h
        header_style(c_h, "#34495E", size=8)

        c_v = ws[f"{col_letter}{row+1}"]
        c_v.value = v
        data_style(c_v, bold=True, align="center", size=10)
        c_v.border = thin

    row += 3

    # ── SECTIONS ──────────────────────────────────────────────────────────────────
    headers = ["S.No", "Item Name", "Technical Description", "Unit", "Quantity", "Calculation Basis"]

    for section in boq_sections:
        sec_id   = section.get("section_id", "")
        sec_name = section.get("section_name", "")
        items    = section.get("items", [])

        # Section Header
        ws.merge_cells(f"A{row}:F{row}")
        cell = ws[f"A{row}"]
        cell.value = f"SECTION {sec_id}: {sec_name.upper()}"
        header_style(cell, "#16A085", size=10)
        ws.row_dimensions[row].height = 22
        row += 1

        # Table Column Headers
        for col_idx, h in enumerate(headers, start=1):
            c = ws[f"{get_column_letter(col_idx)}{row}"]
            c.value = h
            header_style(c, "#2C3E50", size=9)
        ws.row_dimensions[row].height = 20
        row += 1

        # Data Rows
        for item in items:
            row_vals = [
                item.get("sno", ""),
                item.get("item", ""),
                item.get("description", ""),
                item.get("unit", ""),
                item.get("quantity", 0),
                item.get("calculation_basis", ""),
            ]
            aligns = ["center", "left", "left", "center", "right", "left"]
            for col_idx, (v, align) in enumerate(zip(row_vals, aligns), start=1):
                c = ws[f"{get_column_letter(col_idx)}{row}"]
                c.value = v
                data_style(c, align=align)
                c.border = thin
                if align == "right" and isinstance(v, (int, float)):
                    c.number_format = "#,##0.0"
            ws.row_dimensions[row].height = 28
            row += 1

        row += 1  # Blank row between sections

    # ── FOOTER / DISCLAIMER ───────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws[f"A{row}"]
    cell.value = (
        f"DISCLAIMER: This BOQ is automatically generated based on spatial analysis and complies with {std_notes}. "
        "All quantities are preliminary estimates and must be verified by a certified engineer before commercial use."
    )
    cell.font = Font(italic=True, size=8, color="7F8C8D")
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def generate_csv(boq_sections: list) -> str:
    """Generate CSV string from BOQ sections."""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow([
        "Section ID", "Section Name", "S.No", "Item Name",
        "Technical Description", "Unit", "Quantity", "Calculation Basis"
    ])
    
    for section in boq_sections:
        sec_id = section.get("section_id", "")
        sec_name = section.get("section_name", "")
        for item in section.get("items", []):
            writer.writerow([
                sec_id,
                sec_name,
                item.get("sno", ""),
                item.get("item", ""),
                item.get("description", ""),
                item.get("unit", ""),
                item.get("quantity", 0),
                item.get("calculation_basis", ""),
            ])
            
    return output.getvalue()
